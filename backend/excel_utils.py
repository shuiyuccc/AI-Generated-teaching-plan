"""
Excel 工具模块（桩模块 - 用于开发调试）
提供教案模板下载和 Excel 解析功能
"""
import io
import logging

logger = logging.getLogger(__name__)


def create_lesson_plan_template():
    """
    创建教案填写模板 Excel

    Returns:
        BytesIO 对象，包含 Excel 文件内容
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教案批量填写"

        # 固定信息区
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        label_font = Font(name='微软雅黑', size=11, bold=True)
        normal_font = Font(name='微软雅黑', size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 标题
        ws.merge_cells('A1:H1')
        ws['A1'] = '教案批量生成 - Excel 模板'
        ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='007AFF')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36

        # 固定信息
        ws.merge_cells('A2:H2')
        ws['A2'] = '▶ 固定信息（所有课时共用）'
        ws['A2'].font = Font(name='微软雅黑', size=12, bold=True, color='FF9500')

        fixed_fields = [
            ('院系', '智能装备学院'), ('授课班级', '电气自动化（2）班'),
            ('专业名称', '电气自动化'), ('课程名称', '电子焊接'),
            ('授课教师', '张老师')
        ]
        for i, (label, example) in enumerate(fixed_fields):
            row = 3 + i
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=example).font = normal_font
            ws.cell(row=row, column=1).fill = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')

        # 课时信息
        ws.merge_cells('A9:H9')
        ws['A9'] = '▶ 课时信息（每行一个课时）'
        ws['A9'].font = Font(name='微软雅黑', size=12, bold=True, color='FF9500')

        headers = ['序号', '课题名称', '授课地点', '授课时间', '授课学时', '授课类型', '参考资料路径(选填)', '备注(选填)']
        header_row = 10
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 示例数据
        examples = [
            [1, '焊接5步法', '焊接实训室', '2026年2月', '3学时', '理实一体化', '', ''],
            [2, '电子元器件识别', '电子实训室', '2026年3月', '2学时', '理论课', '', ''],
            [3, '电路板焊接实操', '焊接实训室', '2026年3月', '4学时', '实践课', '', ''],
        ]
        for i, row_data in enumerate(examples):
            row = header_row + 1 + i
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = normal_font
                cell.border = thin_border
                if col == 1:
                    cell.alignment = Alignment(horizontal='center')

        # 列宽
        col_widths = [6, 25, 18, 16, 12, 14, 30, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except ImportError:
        logger.error("openpyxl 未安装")
        raise


def parse_lesson_plan_excel(file_path):
    """
    解析教案填写 Excel

    Args:
        file_path: Excel 文件路径

    Returns:
        dict: {
            'success': True/False,
            'fixed_info': {...},
            'lessons': [{...}, ...],
            'document_paths': {...},
            'error': '错误信息'
        }
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        fixed_info = {}
        lessons = []
        document_paths = {}

        # 读取固定信息（行3-7）
        fixed_field_names = ['院系', '授课班级', '专业名称', '课程名称', '授课教师']
        for i, name in enumerate(fixed_field_names):
            row = 3 + i
            label = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            if label and value:
                fixed_info[name] = str(value).strip()

        # 读取课时信息（从第11行开始）
        for row in range(11, ws.max_row + 1):
            seq = ws.cell(row=row, column=1).value
            if seq is None:
                continue

            lesson = {
                'id': str(int(seq)) if isinstance(seq, (int, float)) else str(seq),
                '课题名称': str(ws.cell(row=row, column=2).value or '').strip(),
                '授课地点': str(ws.cell(row=row, column=3).value or '').strip(),
                '授课时间': str(ws.cell(row=row, column=4).value or '').strip(),
                '授课学时': str(ws.cell(row=row, column=5).value or '').strip(),
                '授课类型': str(ws.cell(row=row, column=6).value or '').strip(),
            }

            doc_path = str(ws.cell(row=row, column=7).value or '').strip()
            if doc_path:
                document_paths[str(int(seq))] = doc_path

            if lesson['课题名称']:
                lessons.append(lesson)

        if not lessons:
            return {'success': False, 'error': '未找到课时信息，请确保从第11行开始填写'}

        return {
            'success': True,
            'fixed_info': fixed_info,
            'lessons': lessons,
            'document_paths': document_paths
        }

    except ImportError:
        return {'success': False, 'error': 'openpyxl 未安装'}
    except Exception as e:
        logger.error(f"Excel 解析失败: {e}")
        return {'success': False, 'error': f'解析失败: {str(e)}'}
